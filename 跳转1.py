import sys
import os
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QHBoxLayout, QFileDialog, QProgressBar, QLabel, QLineEdit
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap, QIcon
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

# 新的界面：图片插入Excel程序
class InsertImagesToExcelWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Picture to Excel Program')

        # 设置标签和输入框
        self.account_label = QLabel('Image folder path:')
        self.account_input = QLineEdit(self)
        self.account_input.setPlaceholderText("Choose the image folder...")

        self.password_label = QLabel('Target Excel output folder:')
        self.password_input = QLineEdit(self)
        self.password_input.setPlaceholderText("Choose the output folder...")

        # 文件夹选择按钮
        self.account_button = QPushButton('Choose Image Folder', self)
        self.account_button.clicked.connect(self.choose_image_folder)

        self.password_button = QPushButton('Choose Output Folder', self)
        self.password_button.clicked.connect(self.choose_output_folder)

        # 进度条
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setAlignment(Qt.AlignCenter)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)

        # 启动按钮
        self.start_button = QPushButton('Start Generating', self)
        self.start_button.clicked.connect(self.save_credentials)

        # 布局
        layout = QVBoxLayout()
        layout.addWidget(self.account_label)
        layout.addWidget(self.account_input)
        layout.addWidget(self.account_button)
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(self.password_button)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.start_button)

        self.setLayout(layout)

    def choose_image_folder(self):
        folder = QFileDialog.getExistingDirectory(self, 'Select Image Folder')
        if folder:
            self.account_input.setText(folder)

    def choose_output_folder(self):
        folder = QFileDialog.getExistingDirectory(self, 'Select Output Folder')
        if folder:
            self.password_input.setText(folder)

    def save_credentials(self):
        image_folder = self.account_input.text()
        output_folder = self.password_input.text()

        if not image_folder or not output_folder:
            self.show_error("Please specify both input and output folders.")
            return

        # 启动文件生成进程
        self.progress_bar.setValue(0)  # Reset progress bar
        self.start_button.setEnabled(False)

        # 调用函数将图片插入Excel文件中
        try:
            insert_images_to_excel(image_folder, output_folder, self.update_progress)
            self.show_success("Task completed successfully!")
        except Exception as e:
            self.show_error(f"An error occurred: {str(e)}")

        self.start_button.setEnabled(True)

    def update_progress(self, value):
        self.progress_bar.setValue(int(value))

    def show_error(self, message):
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat(f"Error: {message}")

    def show_success(self, message):
        self.progress_bar.setValue(100)
        self.progress_bar.setFormat(f"Success: {message}")


def insert_images_to_excel(image_folder, output_folder, progress_callback):
    total_images = sum([len(files) for _, _, files in os.walk(image_folder) if any(f.endswith(('.png', '.jpg', '.jpeg', 'JPG')) for f in files)])
    current_image = 0

    # 遍历指定文件夹中的所有子文件夹
    for subfolder in os.listdir(image_folder):
        subfolder_path = os.path.join(image_folder, subfolder)
        if os.path.isdir(subfolder_path):
            # 创建一个新的工作簿
            wb = Workbook()
            ws = wb.active

            # 获取子文件夹名称作为Excel文件的名称
            excel_file = f"{output_folder}/{subfolder}.xlsx"

            # 初始化行和列的计数器
            row = 1
            col = 1

            # 遍历子文件夹中的所有图片文件
            for image_file in sorted(os.listdir(subfolder_path)):
                if image_file.endswith(('.png', '.jpg', '.jpeg', 'JPG')):
                    img_path = os.path.join(subfolder_path, image_file)
                    img = Image.open(img_path)

                    # 将图片大小调整为高2000，宽1500 cm（转换为像素）
                    img = img.resize((2000, 1500))  # 宽度和高度需要转换为像素
                    img.save(img_path)

                    # 将图片插入到Excel表格中
                    img = XLImage(img_path)
                    img.width = 903 // 6 * 2.3
                    img.height = 677 // 6 * 2.3

                    ws.column_dimensions[get_column_letter(col)].width = img.width // 7.3
                    ws.row_dimensions[row].height = img.height * 0.8

                    ws.add_image(img, f"{get_column_letter(col)}{row}")

                    # 更新行和列的计数器
                    col += 1
                    if col > 5:
                        col = 1
                        row += 1

                    # 更新进度条
                    current_image += 1
                    progress_callback(current_image / total_images * 100)

            # 保存Excel文件
            wb.save(excel_file)


# 主窗口
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Main Window')

        # 设置按钮
        self.btn_convert_to_word = QPushButton('图片文件夹转换为Word')
        self.btn_convert_to_excel = QPushButton('图片文件夹转换为Excel')
        self.btn_compress_image = QPushButton('压缩图片')
        self.btn_crop_image = QPushButton('裁剪图片')
        self.btn_ratio_converter = QPushButton('比例转换器')

        # 连接按钮事件
        self.btn_convert_to_excel.clicked.connect(self.open_insert_images_to_excel_window)

        # 创建布局
        layout = QVBoxLayout()
        layout.addWidget(self.btn_convert_to_word)
        layout.addWidget(self.btn_convert_to_excel)
        layout.addWidget(self.btn_compress_image)
        layout.addWidget(self.btn_crop_image)
        layout.addWidget(self.btn_ratio_converter)

        self.setLayout(layout)

        # 设置窗口图标
        self.setWindowIcon(QIcon("icon.png"))

    def open_insert_images_to_excel_window(self):
        self.insert_images_to_excel_window = InsertImagesToExcelWindow()
        self.insert_images_to_excel_window.show()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
